from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Literal

import psycopg2
from openpyxl import load_workbook

KNOWN_PARTY_ORDERS: dict[str, int] = {
    "PCP-PEV": 10,
    "PCTP/MRPP": 15,
    "B.E.": 20,
    "L": 30,
    "PAN": 35,
    "PS": 40,
    "JPP": 45,
    "MPT": 50,
    "PDR": 52,
    "VP": 55,
    "IL": 60,
    "R.I.R.": 62,
    "ADN": 65,
    "NC": 67,
    "ND": 68,
    "PPD/PSD": 70,
    "A": 72,
    "CDS-PP": 80,
    "PPM": 85,
    "CH": 90,
    "MAS": 18,
    "E": 92,
    "PLS": 57,
    "PNR": 88,
    "PPV/CDC": 82,
    "PTP": 42,
    "PURP": 43,
}

KNOWN_PARTY_ALIASES: dict[str, tuple[str, ...]] = {
    "PCP-PEV": ("PCP-PEV", "CDU"),
    "B.E.": ("B.E.", "BE"),
    "R.I.R.": ("R.I.R.", "RIR"),
    "VP": ("VP", "VOLT"),
    "PPD/PSD": ("PPD/PSD", "PSD"),
    "CDS-PP": ("CDS-PP", "CDS/PP", "CDS"),
    "NC": ("NC",),
    "ND": ("ND",),
}

KNOWN_PARTIES = set(KNOWN_PARTY_ORDERS)

# Headers that may appear in CNE files. AD is intentionally here even though it is
# a coalition, because some source files expose it like a normal result column.
PARTY_HEADERS = KNOWN_PARTIES | {"AD"}

EXPLICIT_COALITION_SIGLAS = {"AD", "PCP-PEV", "CDU"}
LEGISLATIVE_AD_ELECTIONS = {"LEGISLATIVAS_2024", "LEGISLATIVAS_2025"}
LEGISLATIVE_AD_VARIANTS = {
    "AD",
    "PPD/PSD.CDS-PP",
    "PPD/PSD.CDS-PP.PPM",
    "PPD/PSD.CDS-PP.PPM.A",
}

GCE_PREFIXES = ("GCE-", "MOV", "M.A")

LEGISLATIVE_OFFICE_CODE = "AR"


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)

    s = str(value).strip().replace(" ", "")
    if not s or s == "-":
        return None

    try:
        return int(float(s.replace(",", ".")))
    except ValueError:
        return None


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def normalize_header_text(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper().strip()
    return re.sub(r"\s+", " ", text)


def normalize_sigla(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None

    sigla = re.sub(r"\s+", "", text.upper())

    # Fix known typo from source files.
    sigla = sigla.replace("PPD/PDS", "PPD/PSD")

    if sigla in {"BE", "B.E"}:
        return "B.E."

    if sigla in {"RIR", "R.I.R"}:
        return "R.I.R."

    if sigla in {"VOLT", "VOLT."}:
        return "VP"

    return sigla


def canonical_sigla_for_election(sigla: str, election_code: str) -> str:
    normalized = normalize_sigla(sigla) or sigla

    # Legislative AD must be aggregated only for the elections where AD was the
    # real national candidacy. Do not apply this to LEGISLATIVAS_2022 or to
    # autárquicas, because there PSD/CDS/PPM combinations are real separate lists.
    if (
        election_code in LEGISLATIVE_AD_ELECTIONS
        and normalized in LEGISLATIVE_AD_VARIANTS
    ):
        return "AD"

    return normalized


def normalize_code(value: Any) -> str | None:
    if value is None or value == "":
        return None

    s = str(value).strip()
    if re.fullmatch(r"\d+(\.0)?", s):
        s = str(int(float(s)))
    if len(s) < 6:
        s = s.zfill(6)
    return s


def canonical_territory_code(
    raw_code: str, office_code: str, freguesia: str | None
) -> str:
    code = raw_code.zfill(6)
    if office_code == "AF":
        return code
    if office_code in {"CM", "AM"}:
        if code.endswith("0000"):
            return code[:2]
        return code[:4]
    if office_code == LEGISLATIVE_OFFICE_CODE:
        return code[:2]
    if freguesia and not re.fullmatch(r"\(\d+\)", freguesia.strip()):
        return code
    if code.endswith("0000"):
        return code[:2]
    return code[:4]


def parse_bracket_labels(cell: Any) -> list[str]:
    text = clean_text(cell)
    if not text:
        return []

    labels: list[str] = []

    for match in re.findall(r"\[([^\]]+)\]", text):
        sigla = normalize_sigla(match)
        if sigla and sigla not in labels:
            labels.append(sigla)

    return labels


def _member_pattern(alias: str) -> re.Pattern[str]:
    return re.compile(rf"(^|[.+/\-]){re.escape(alias)}($|[.+/\-])")


PARTY_MEMBER_PATTERNS: list[tuple[str, re.Pattern[str]]] = []
for _member_sigla in KNOWN_PARTY_ORDERS:
    aliases = KNOWN_PARTY_ALIASES.get(_member_sigla, (_member_sigla,))
    alias_group = "|".join(re.escape(alias) for alias in aliases)
    PARTY_MEMBER_PATTERNS.append(
        (_member_sigla, re.compile(rf"(^|[.+/\-])(?:{alias_group})($|[.+/\-])"))
    )


def is_exact_known_party(sigla: str) -> bool:
    normalized = normalize_sigla(sigla) or sigla
    if normalized in KNOWN_PARTIES and normalized not in {"AD", "PCP-PEV"}:
        return True
    return normalized in {"B.E.", "R.I.R.", "VP"}


def detect_known_party_members(sigla: str) -> list[str]:
    normalized = normalize_sigla(sigla) or sigla

    if normalized == "AD":
        return ["PPD/PSD", "CDS-PP", "PPM"]

    members: list[str] = []
    for member_sigla, pattern in PARTY_MEMBER_PATTERNS:
        if pattern.search(normalized) and member_sigla not in members:
            members.append(member_sigla)

    members.sort(key=lambda item: KNOWN_PARTY_ORDERS.get(item, 999))
    return members


def looks_like_coalition_sigla(sigla: str) -> bool:
    normalized = normalize_sigla(sigla) or sigla

    if normalized.startswith(GCE_PREFIXES):
        return False

    if normalized in EXPLICIT_COALITION_SIGLAS:
        return True

    if is_exact_known_party(normalized):
        return False

    return len(detect_known_party_members(normalized)) >= 2


def safe_gce_sigla(sigla: str) -> str:
    normalized = normalize_sigla(sigla) or sigla

    if normalized in PARTY_HEADERS:
        return f"GCE-{normalized}"

    return normalized


def entity_type_for_sigla(sigla: str, source_entity_type: str | None = None) -> str:
    normalized = normalize_sigla(sigla) or sigla

    # When the CNE autárquicas file explicitly puts a label under SIGLAS COLIGAÇÕES
    # or SIGLAS GCE, trust the source. This keeps local labels like O.I., C.R. and
    # F.P. as coalitions even when they cannot be decomposed into national parties.
    if source_entity_type in {"coalition", "gce", "blank", "null", "other"}:
        return source_entity_type

    if normalized in EXPLICIT_COALITION_SIGLAS:
        return "coalition"

    if is_exact_known_party(normalized):
        return "party"

    if normalized.startswith(GCE_PREFIXES):
        return "gce"

    if len(detect_known_party_members(normalized)) >= 2:
        return "coalition"

    return "party"


@dataclass(frozen=True)
class ParsedSeat:
    territory_code: str
    office_code: str
    sigla: str
    entity_type: str
    seats: int


@dataclass(frozen=True)
class ParsedRow:
    row_no: int
    raw_code: str
    territory_code: str
    territory_level: Literal["country", "district", "municipality", "parish"]
    territory_name: str | None
    parent_code: str | None
    concelho: str | None
    freguesia: str | None
    office_code: str
    registered_voters: int
    voters: int
    blank_votes: int
    null_votes: int
    votes: list[tuple[str, str, int, int]]
    raw: dict[str, Any]


class AutarquicasParser:
    FIXED_COLS = 8

    def __init__(self, path: Path, sheet_name: str | None = None):
        if path.suffix.lower() == ".xls":
            raise SystemExit(
                "Wrong file format. Convert .xls to .xlsx first with LibreOffice."
            )

        self.path = path
        self.workbook = load_workbook(path, read_only=True, data_only=True)
        self.sheet = (
            self.workbook[sheet_name]
            if sheet_name
            else self.workbook[self.workbook.sheetnames[0]]
        )
        self.sheet_name = self.sheet.title

        self.header_row = self.detect_header_row()
        self.data_start_row = self.header_row + 1
        self.group_headers = [
            clean_text(c.value) for c in self.sheet[self.header_row - 1]
        ]
        self.headers = [clean_text(c.value) for c in self.sheet[self.header_row]]

        self.last_coalition_labels_ordered: list[str] = []
        self.last_gce_labels_ordered: list[str] = []

        if len(self.headers) >= 8:
            first_three = [normalize_header_text(v) for v in self.headers[:3]]
            if first_three == ["COD", "CONC", "FREG"]:
                self.headers[0:8] = [
                    "CÓD",
                    "CONC",
                    "FREG",
                    "ÓRG",
                    "inscritos",
                    "votantes",
                    "brancos",
                    "nulos",
                ]

    def detect_header_row(self) -> int:
        for row_no in range(1, 30):
            row_values = [normalize_header_text(c.value) for c in self.sheet[row_no]]
            has_code = any(
                v in {"COD", "COD.", "CODIGO", "CODIGO.", "CÓD", "CÓD."}
                for v in row_values
            )
            has_conc = "CONC" in row_values
            has_freg = "FREG" in row_values
            has_office = any(
                v in {"ORG", "ORG.", "ORGAO", "ORGAO.", "ÓRG", "ÓRG."}
                for v in row_values
            )
            has_registered = any(v in {"INSC", "INSCRITOS"} for v in row_values)
            has_voters = any(v in {"VOT", "VOTANTES"} for v in row_values)

            if has_code and has_office and has_registered and has_voters:
                return row_no
            if has_code and has_conc and has_freg:
                return row_no

        raise RuntimeError(f"Could not detect header row in sheet {self.sheet_name}.")

    def result_columns(self) -> list[tuple[int, str]]:
        result_cols: list[tuple[int, str]] = []
        for idx, header in enumerate(self.headers, start=1):
            if idx <= self.FIXED_COLS or not header:
                continue
            is_placeholder = re.fullmatch(r"\[[A-Z]\]", header) is not None
            is_special_label_col = "SIGLA" in header.upper()
            if is_special_label_col:
                continue
            if header.startswith("[") and header.endswith("]") and not is_placeholder:
                continue
            result_cols.append((idx, header))
        return result_cols

    def is_column_group(self, idx: int, wanted: str) -> bool:
        header = self.headers[idx] if idx < len(self.headers) else None
        group_header = (
            self.group_headers[idx] if idx < len(self.group_headers) else None
        )
        header_norm = clean_text(header)
        group_norm = clean_text(group_header)
        if wanted == "coalition":
            return header_norm == "[SIGLA COL]" or group_norm == "SIGLAS COLIGAÇÕES"
        if wanted == "gce":
            return header_norm == "[SIGLA GCE]" or group_norm == "SIGLAS GCE"
        return False

    def row_coalition_labels(self, values: list[Any]) -> list[str]:
        labels: list[str] = []
        for idx, value in enumerate(values):
            header = self.headers[idx] if idx < len(self.headers) else None
            header_is_placeholder = (
                header is not None and re.fullmatch(r"\[[A-Z]\]", header) is not None
            )
            if self.is_column_group(idx, "coalition"):
                for label in parse_bracket_labels(value):
                    if label not in labels:
                        labels.append(label)
            elif self.is_column_group(idx, "gce"):
                continue
            elif idx >= self.FIXED_COLS and not header_is_placeholder:
                for label in parse_bracket_labels(value):
                    if looks_like_coalition_sigla(label) and label not in labels:
                        labels.append(label)
        return labels

    def row_gce_labels(self, values: list[Any]) -> list[str]:
        labels: list[str] = []
        for idx, value in enumerate(values):
            header = self.headers[idx] if idx < len(self.headers) else None
            header_is_placeholder = (
                header is not None and re.fullmatch(r"\[[A-Z]\]", header) is not None
            )
            if self.is_column_group(idx, "gce"):
                for label in parse_bracket_labels(value):
                    label = safe_gce_sigla(label)
                    if label not in labels:
                        labels.append(label)
            elif self.is_column_group(idx, "coalition"):
                continue
            elif idx >= self.FIXED_COLS and not header_is_placeholder:
                for label in parse_bracket_labels(value):
                    if not looks_like_coalition_sigla(label):
                        label = safe_gce_sigla(label)
                        if label not in labels:
                            labels.append(label)
        return labels

    def parse_rows(self) -> Iterable[ParsedRow]:
        result_columns = self.result_columns()

        for row_no, row in enumerate(
            self.sheet.iter_rows(min_row=self.data_start_row, values_only=True),
            start=self.data_start_row,
        ):
            values = list(row)
            raw_code = normalize_code(values[0] if len(values) > 0 else None)
            concelho = clean_text(values[1] if len(values) > 1 else None)
            freguesia = clean_text(values[2] if len(values) > 2 else None)
            office_code = clean_text(values[3] if len(values) > 3 else None)
            if not raw_code or not office_code:
                continue

            registered = as_int(values[4] if len(values) > 4 else None)
            voters = as_int(values[5] if len(values) > 5 else None)
            blank = as_int(values[6] if len(values) > 6 else None)
            null = as_int(values[7] if len(values) > 7 else None)
            if registered is None or voters is None or blank is None or null is None:
                continue

            territory_code = canonical_territory_code(raw_code, office_code, freguesia)
            raw_6 = raw_code.zfill(6)

            if office_code == "AF":
                territory_level: Literal[
                    "country", "district", "municipality", "parish"
                ] = "parish"
                territory_name = freguesia
                parent_code = raw_6[:4]
            elif office_code in {"CM", "AM"}:
                if territory_code == raw_6[:2]:
                    territory_level = "district"
                    territory_name = f"Distrito {territory_code}"
                    parent_code = "PT"
                else:
                    territory_level = "municipality"
                    territory_name = concelho
                    parent_code = raw_6[:2]
            else:
                territory_level = "district"
                territory_name = concelho or freguesia or territory_code
                parent_code = "PT"

            row_coalition_labels = self.row_coalition_labels(values)
            row_gce_labels = self.row_gce_labels(values)
            if row_coalition_labels:
                self.last_coalition_labels_ordered = row_coalition_labels
            if row_gce_labels:
                self.last_gce_labels_ordered = row_gce_labels

            coalition_labels_ordered = (
                row_coalition_labels or self.last_coalition_labels_ordered
            )
            gce_labels_ordered = row_gce_labels or self.last_gce_labels_ordered

            votes: list[tuple[str, str, int, int]] = []
            coalition_votes_seen = 0
            gce_votes_seen = 0

            for display_order, (col_idx, raw_sigla) in enumerate(
                result_columns, start=1
            ):
                vote_count = as_int(
                    values[col_idx - 1] if col_idx - 1 < len(values) else None
                )
                if vote_count is None:
                    continue

                if re.fullmatch(r"\[[A-Z]\]", raw_sigla):
                    placeholder_letter = raw_sigla.strip("[]")

                    if placeholder_letter in {"A", "B", "C"}:
                        if coalition_votes_seen < len(coalition_labels_ordered):
                            sigla = coalition_labels_ordered[coalition_votes_seen]
                        else:
                            if vote_count == 0:
                                coalition_votes_seen += 1
                                continue
                            sigla = f"SOURCE-COL-{placeholder_letter}-{territory_code}-{office_code}"

                        coalition_votes_seen += 1
                        entity_type = "coalition"

                    elif placeholder_letter in {"D", "E", "F", "G"}:
                        if gce_votes_seen < len(gce_labels_ordered):
                            sigla = gce_labels_ordered[gce_votes_seen]
                        else:
                            if vote_count == 0:
                                gce_votes_seen += 1
                                continue
                            sigla = f"SOURCE-GCE-{placeholder_letter}-{territory_code}-{office_code}"

                        gce_votes_seen += 1
                        entity_type = "gce"

                    else:
                        if vote_count == 0:
                            continue
                        sigla = f"SOURCE-OTHER-{placeholder_letter}-{territory_code}-{office_code}"
                        entity_type = "other"

                else:
                    sigla = normalize_sigla(raw_sigla)
                    if not sigla:
                        continue
                    entity_type = entity_type_for_sigla(sigla)

                sigla = normalize_sigla(sigla)
                if not sigla:
                    continue

                entity_type = entity_type_for_sigla(
                    sigla, source_entity_type=entity_type
                )

                votes.append((sigla, entity_type, vote_count, display_order))

            raw = {
                str(self.headers[i] or f"col_{i + 1}"): values[i]
                for i in range(min(len(values), len(self.headers)))
            }

            yield ParsedRow(
                row_no=row_no,
                raw_code=raw_code,
                territory_code=territory_code,
                territory_level=territory_level,
                territory_name=territory_name,
                parent_code=parent_code,
                concelho=concelho,
                freguesia=freguesia,
                office_code=office_code,
                registered_voters=registered,
                voters=voters,
                blank_votes=blank,
                null_votes=null,
                votes=votes,
                raw=raw,
            )


class LegislativasParser:
    """Parser for CNE AR matrix workbooks like 2022ar_quadro_resultados.xlsx."""

    def __init__(self, path: Path, sheet_name: str | None = None):
        if path.suffix.lower() == ".xls":
            raise SystemExit(
                "Wrong file format. Convert .xls to .xlsx first with LibreOffice."
            )

        self.path = path
        self.workbook = load_workbook(path, read_only=True, data_only=True)
        self.sheet = (
            self.workbook[sheet_name]
            if sheet_name
            else self.workbook[self.workbook.sheetnames[0]]
        )
        self.sheet_name = self.sheet.title
        self.circle_header_row = self.detect_circle_header_row()
        self.circle_name_row = self.circle_header_row + 1

    def detect_circle_header_row(self) -> int:
        for row_no in range(1, min(self.sheet.max_row, 20) + 1):
            row_values = [normalize_header_text(c.value) for c in self.sheet[row_no]]
            if "CIRCULO" in row_values or "CÍRCULO" in row_values:
                return row_no
        raise RuntimeError(
            f"Could not detect legislative circle header row in sheet {self.sheet_name}."
        )

    def circle_columns(self) -> list[tuple[int, str, str]]:
        cols: list[tuple[int, str, str]] = []
        for col_idx in range(1, self.sheet.max_column + 1):
            circle_no = self.sheet.cell(self.circle_header_row, col_idx).value
            circle_name = clean_text(
                self.sheet.cell(self.circle_name_row, col_idx).value
            )
            if circle_name is None:
                continue
            if normalize_header_text(circle_name) == "TOTAL":
                continue

            circle_code = as_int(circle_no)
            if circle_code is None:
                continue
            cols.append((col_idx, f"{circle_code:02d}", circle_name))

        if not cols:
            raise RuntimeError("No legislative result-circle columns found.")
        return cols

    def metric_value_by_labels(
        self,
        labels: list[str],
        circle_cols: list[tuple[int, str, str]],
    ) -> dict[str, int]:
        wanted = {normalize_header_text(label) for label in labels}

        for row_no in range(1, self.sheet.max_row + 1):
            if normalize_header_text(self.sheet.cell(row_no, 1).value) not in wanted:
                continue

            out: dict[str, int] = {}

            for col_idx, code, _name in circle_cols:
                value = as_int(self.sheet.cell(row_no, col_idx).value)

                if value is None:
                    raise RuntimeError(
                        f"Missing metric {labels!r} for circle {code} "
                        f"at row {row_no}, column {col_idx}."
                    )

                out[code] = value

            return out

        raise RuntimeError(f"Could not find required metric row among {labels!r}.")

    def parse_rows(self) -> Iterable[ParsedRow]:
        circle_cols = self.circle_columns()
        registered_by_code = self.metric_value_by_labels(["Inscritos"], circle_cols)
        voters_by_code = self.metric_value_by_labels(
            ["Votantes", "Votantes (VTT)"], circle_cols
        )
        blank_by_code = self.metric_value_by_labels(["Brancos"], circle_cols)
        null_by_code = self.metric_value_by_labels(["Nulos"], circle_cols)

        party_vote_rows: list[tuple[int, str]] = []

        non_party_rows = {
            "INSCRITOS",
            "VOTANTES",
            "VOTANTES(VTT)",
            "ABSTENCAO",
            "ABSTENÇÃO",
            "BRANCOS",
            "NULOS",
            "VOTOSVAL.EXP.(VVE)",
            "VOTOSVALEXP.(VVE)",
            "VOTOSVALIDAMENTEEXPRESSOS",
        }

        for row_no in range(1, self.sheet.max_row + 1):
            sigla = normalize_sigla(self.sheet.cell(row_no, 1).value)
            metric = normalize_header_text(self.sheet.cell(row_no, 2).value)

            if not sigla or metric != "NUMERO":
                continue

            if sigla in non_party_rows:
                continue

            party_vote_rows.append((row_no, sigla))

        if not party_vote_rows:
            raise RuntimeError("No legislative party vote rows found.")

        for territory_order, (col_idx, territory_code, territory_name) in enumerate(
            circle_cols, start=1
        ):
            votes: list[tuple[str, str, int, int]] = []
            for display_order, (vote_row_no, sigla) in enumerate(
                party_vote_rows, start=1
            ):
                vote_count = as_int(self.sheet.cell(vote_row_no, col_idx).value)
                if vote_count is None:
                    continue

                sigla = normalize_sigla(sigla)
                if not sigla:
                    continue

                votes.append(
                    (sigla, entity_type_for_sigla(sigla), vote_count, display_order)
                )

            yield ParsedRow(
                row_no=territory_order,
                raw_code=territory_code,
                territory_code=territory_code,
                territory_level="district",
                territory_name=territory_name,
                parent_code="PT",
                concelho=None,
                freguesia=None,
                office_code=LEGISLATIVE_OFFICE_CODE,
                registered_voters=registered_by_code[territory_code],
                voters=voters_by_code[territory_code],
                blank_votes=blank_by_code[territory_code],
                null_votes=null_by_code[territory_code],
                votes=votes,
                raw={"circle_code": territory_code, "circle_name": territory_name},
            )

    def parse_seats(self) -> Iterable[ParsedSeat]:
        circle_cols = self.circle_columns()
        party_vote_rows: list[tuple[int, str]] = []

        non_party_rows = {
            "INSCRITOS",
            "VOTANTES",
            "VOTANTES(VTT)",
            "ABSTENCAO",
            "ABSTENÇÃO",
            "BRANCOS",
            "NULOS",
            "VOTOSVAL.EXP.(VVE)",
            "VOTOSVALEXP.(VVE)",
            "VOTOSVALIDAMENTEEXPRESSOS",
        }

        for row_no in range(1, self.sheet.max_row + 1):
            sigla = normalize_sigla(self.sheet.cell(row_no, 1).value)
            metric = normalize_header_text(self.sheet.cell(row_no, 2).value)

            if not sigla or metric != "NUMERO":
                continue

            if sigla in non_party_rows:
                continue

            party_vote_rows.append((row_no, sigla))

        for vote_row_no, sigla in party_vote_rows:
            seat_row_no = vote_row_no + 2
            if normalize_header_text(self.sheet.cell(seat_row_no, 2).value) != "MD":
                continue
            for col_idx, territory_code, _territory_name in circle_cols:
                seat_count = as_int(self.sheet.cell(seat_row_no, col_idx).value)
                if seat_count is None or seat_count <= 0:
                    continue

                sigla = normalize_sigla(sigla)
                if not sigla:
                    continue

                yield ParsedSeat(
                    territory_code=territory_code,
                    office_code=LEGISLATIVE_OFFICE_CODE,
                    sigla=sigla,
                    entity_type=entity_type_for_sigla(sigla),
                    seats=seat_count,
                )


def detect_parser_mode(
    path: Path, sheet_name: str | None
) -> Literal["autarquicas", "legislativas"]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    for row_no in range(1, min(sheet.max_row, 20) + 1):
        row_values = [normalize_header_text(c.value) for c in sheet[row_no]]
        if any(
            v in {"COD", "COD.", "CODIGO", "CODIGO.", "CÓD", "CÓD."} for v in row_values
        ):
            return "autarquicas"
        if "CIRCULO" in row_values or "CÍRCULO" in row_values:
            return "legislativas"

    raise RuntimeError(
        "Could not detect workbook layout. Use --layout autarquicas or --layout legislativas."
    )


def get_one(cur, sql: str, params: tuple[Any, ...]) -> Any:
    cur.execute(sql, params)
    row = cur.fetchone()
    if row is None:
        raise RuntimeError(f"Query returned no rows: {sql}")
    return row[0]


def build_territory_rows(
    parsed_rows: list[ParsedRow],
) -> list[tuple[str, str, str, str | None]]:
    territories: dict[str, tuple[str, str, str, str | None]] = {}
    territories["PT"] = ("PT", "country", "Portugal", None)

    for pr in parsed_rows:
        raw = pr.raw_code.zfill(6)

        if pr.territory_level == "district":
            territories[pr.territory_code] = (
                pr.territory_code,
                "district",
                pr.territory_name or f"Círculo {pr.territory_code}",
                pr.parent_code or "PT",
            )
            continue

        district_code = raw[:2]
        municipality_code = raw[:4]
        parish_code = raw

        territories.setdefault(
            district_code,
            (district_code, "district", f"Distrito {district_code}", "PT"),
        )

        if raw.endswith("0000") or municipality_code.endswith("00"):
            if pr.office_code == "AF" and pr.freguesia:
                territories[parish_code] = (
                    parish_code,
                    "parish",
                    pr.freguesia,
                    district_code,
                )
            continue

        territories[municipality_code] = (
            municipality_code,
            "municipality",
            pr.concelho or municipality_code,
            district_code,
        )
        if pr.office_code == "AF" and pr.freguesia:
            territories[parish_code] = (
                parish_code,
                "parish",
                pr.freguesia,
                municipality_code,
            )

    level_order = {"country": 0, "district": 1, "municipality": 2, "parish": 3}
    return sorted(territories.values(), key=lambda x: (level_order[x[1]], x[0]))


def bulk_upsert_territories(
    cur, territories: list[tuple[str, str, str, str | None]]
) -> None:
    for code, level_code, name, parent_code in territories:
        cur.execute(
            """
            INSERT INTO op.territory (
                level_id, code, name, parent_id, normalized_name, geom, source_table, source_srid
            )
            SELECT
                tl.territory_level_id,
                %s,
                %s,
                parent.territory_id,
                lower(unaccent(%s)),
                NULL,
                'election_etl_fallback',
                NULL
            FROM op.territory_level tl
            LEFT JOIN op.territory parent ON parent.code = %s
            WHERE tl.code = %s
            ON CONFLICT (code)
            DO UPDATE SET
                name = COALESCE(NULLIF(op.territory.name, ''), EXCLUDED.name),
                parent_id = COALESCE(op.territory.parent_id, EXCLUDED.parent_id),
                normalized_name = COALESCE(op.territory.normalized_name, EXCLUDED.normalized_name),
                updated_at = now()
            """,
            (code, name or code, name or code, parent_code, level_code),
        )


def save_official_seat(
    cur,
    election_code: str,
    office_code: str,
    territory_code: str,
    sigla: str,
    entity_type: str,
    seats: int,
) -> None:
    sigla = canonical_sigla_for_election(sigla, election_code)
    entity_type = entity_type_for_sigla(sigla)

    cur.execute(
        """
        WITH ctx AS (
            SELECT
                e.election_id,
                o.office_id,
                t.territory_id,
                op.save_political_entity(%s, %s) AS political_entity_id
            FROM op.election e
            JOIN op.office o ON o.code = %s
            JOIN op.territory t ON t.code = %s
            WHERE e.code = %s
        ), candidate AS (
            INSERT INTO op.candidacy (
                election_id, office_id, territory_id, political_entity_id, source_label
            )
            SELECT election_id, office_id, territory_id, political_entity_id, %s
            FROM ctx
            ON CONFLICT (election_id, office_id, territory_id, political_entity_id)
            DO UPDATE SET source_label = EXCLUDED.source_label
            RETURNING candidacy_id, election_id, office_id, territory_id
        )
        INSERT INTO op.seat_result (
            election_id, office_id, territory_id, candidacy_id, seats, method, updated_at
        )
        SELECT election_id, office_id, territory_id, candidacy_id, %s, 'official', now()
        FROM candidate
        ON CONFLICT (
            election_id,
            office_id,
            territory_id,
            candidacy_id,
            method
        )
        DO UPDATE SET
            seats = EXCLUDED.seats,
            updated_at = now()
        """,
        (sigla, entity_type, office_code, territory_code, election_code, sigla, seats),
    )


def save_official_seat_counts(cur, election_code: str, seats: list[ParsedSeat]) -> int:
    totals: dict[tuple[str, str], int] = {}
    for seat in seats:
        totals[(seat.office_code, seat.territory_code)] = (
            totals.get((seat.office_code, seat.territory_code), 0) + seat.seats
        )

    for (office_code, territory_code), total_seats in totals.items():
        cur.execute(
            """
            INSERT INTO op.seat_count (election_id, office_id, territory_id, seats, source)
            SELECT e.election_id, o.office_id, t.territory_id, %s, 'official'
            FROM op.election e
            JOIN op.office o ON o.code = %s
            JOIN op.territory t ON t.code = %s
            WHERE e.code = %s
            ON CONFLICT (election_id, office_id, territory_id)
            DO UPDATE SET seats = EXCLUDED.seats, source = EXCLUDED.source
            """,
            (total_seats, office_code, territory_code, election_code),
        )

    return len(totals)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--database-url", default=os.getenv("DATABASE_URL"))
    ap.add_argument("--file", required=True, type=Path)
    ap.add_argument("--sheet")
    ap.add_argument("--election-code", default="AUTARQUICAS_2021")
    ap.add_argument(
        "--layout", choices=["auto", "autarquicas", "legislativas"], default="auto"
    )
    ap.add_argument("--skip-refresh", action="store_true")
    args = ap.parse_args()

    if not args.database_url:
        raise SystemExit(
            "DATABASE_URL missing. Set it in .env/export it or pass --database-url."
        )

    layout = (
        detect_parser_mode(args.file, args.sheet)
        if args.layout == "auto"
        else args.layout
    )
    parser: AutarquicasParser | LegislativasParser
    if layout == "legislativas":
        parser = LegislativasParser(args.file, args.sheet)
    else:
        parser = AutarquicasParser(args.file, args.sheet)

    file_hash = sha256_file(args.file)
    parsed_rows = list(parser.parse_rows())
    parsed_seats = (
        list(parser.parse_seats()) if isinstance(parser, LegislativasParser) else []
    )

    conn = psycopg2.connect(args.database_url)
    conn.autocommit = False

    rows_seen = 0
    rows_loaded = 0
    vote_cells = 0
    seat_rows_loaded = 0
    seat_contexts_loaded = 0

    with conn, conn.cursor() as cur:
        election_id = get_one(
            cur,
            "SELECT election_id FROM op.election WHERE code = %s",
            (args.election_code,),
        )

        import_file_id = get_one(
            cur,
            """
            INSERT INTO op.import_file (election_id, file_path, sheet_name, file_hash)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (election_id, file_hash, sheet_name)
            DO UPDATE SET imported_at = now()
            RETURNING import_file_id
            """,
            (election_id, str(args.file), parser.sheet_name, file_hash),
        )

        territories = build_territory_rows(parsed_rows)
        bulk_upsert_territories(cur, territories)

        for pr in parsed_rows:
            rows_seen += 1
            cur.execute(
                "SELECT op.save_turnout_result(%s,%s,%s,%s,%s,%s,%s,%s)",
                (
                    args.election_code,
                    pr.office_code,
                    pr.territory_code,
                    pr.registered_voters,
                    pr.voters,
                    pr.blank_votes,
                    pr.null_votes,
                    import_file_id,
                ),
            )

            for sigla, entity_type, votes, display_order in pr.votes:
                sigla = canonical_sigla_for_election(sigla, args.election_code)
                entity_type = entity_type_for_sigla(
                    sigla, source_entity_type=entity_type
                )

                cur.execute(
                    "SELECT op.save_candidacy_vote_result(%s,%s,%s,%s,%s,%s,%s,%s)",
                    (
                        args.election_code,
                        pr.office_code,
                        pr.territory_code,
                        sigla,
                        entity_type,
                        votes,
                        display_order,
                        import_file_id,
                    ),
                )
                vote_cells += 1
            rows_loaded += 1

        if parsed_seats:
            seat_contexts_loaded = save_official_seat_counts(
                cur, args.election_code, parsed_seats
            )
            for seat in parsed_seats:
                save_official_seat(
                    cur,
                    args.election_code,
                    seat.office_code,
                    seat.territory_code,
                    seat.sigla,
                    seat.entity_type,
                    seat.seats,
                )
                seat_rows_loaded += 1

        cur.execute("CALL op.populate_seat_count();")
        cur.execute("CALL op.calculate_seat_results();")
        if not args.skip_refresh:
            cur.execute("CALL wh.refresh();")

    conn.close()

    print(
        json.dumps(
            {
                "file": str(args.file),
                "sheet": parser.sheet_name,
                "layout": layout,
                "election_code": args.election_code,
                "territories_preloaded": len(territories),
                "rows_seen": rows_seen,
                "rows_loaded": rows_loaded,
                "vote_cells_loaded": vote_cells,
                "official_seat_contexts_loaded": seat_contexts_loaded,
                "official_seat_rows_loaded": seat_rows_loaded,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
