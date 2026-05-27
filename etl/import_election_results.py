from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import psycopg2
from openpyxl import load_workbook

OFFICIAL_PARTY_HEADERS = {
    "A",
    "ADN",
    "B.E.",
    "CDS-PP",
    "CH",
    "E",
    "IL",
    "JPP",
    "L",
    "MAS",
    "MPT",
    "NC",
    "ND",
    "PAN",
    "PCTP/MRPP",
    "PDR",
    "PLS",
    "PNR",
    "PPD/PSD",
    "PPM",
    "PPV/CDC",
    "PS",
    "PTP",
    "PURP",
    "R.I.R.",
    "VP",
    "PCP-PEV",
}


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()

    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)

    return h.hexdigest()


def as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None

    if isinstance(value, bool):
        return int(value)

    if isinstance(value, (int, float)):
        return int(value)

    s = str(value).strip().replace(" ", "")

    if not s:
        return None

    try:
        return int(float(s.replace(",", ".")))
    except ValueError:
        return None


def clean_text(value: Any) -> str | None:
    if value is None:
        return None

    s = str(value).strip()
    return s or None


def normalize_header_text(value: Any) -> str:
    text = clean_text(value)

    if not text:
        return ""

    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper().strip()
    text = re.sub(r"\s+", " ", text)

    return text


def normalize_code(value: Any) -> str | None:
    if value is None or value == "":
        return None

    s = str(value).strip()

    if re.fullmatch(r"\d+(\.0)?", s):
        s = str(int(float(s)))

    if len(s) < 6:
        s = s.zfill(6)

    return s


def canonical_territory_code(
    raw_code: str,
    office_code: str,
    freguesia: str | None,
) -> str:
    code = raw_code.zfill(6)

    # AF rows are parish-scope.
    # Keep the exact source code, including election-only xxxx00 placeholders.
    if office_code == "AF":
        return code

    # CM/AM rows are municipality-scope, even when old CNE files put
    # footnote-like values such as (1), (2), (8) in the FREG column.
    if office_code in {"CM", "AM"}:
        if code.endswith("0000"):
            return code[:2]

        return code[:4]

    # Fallback for future offices.
    if freguesia and not re.fullmatch(r"\(\d+\)", freguesia.strip()):
        return code

    if code.endswith("0000"):
        return code[:2]

    return code[:4]


def parse_bracket_labels(cell: Any) -> list[str]:
    text = clean_text(cell)

    if not text:
        return []

    return [m.strip() for m in re.findall(r"\[([^\]]+)\]", text) if m.strip()]


def looks_like_coalition_sigla(sigla: str) -> bool:
    return "/" in sigla or "." in sigla


def safe_gce_sigla(sigla: str) -> str:
    if sigla in OFFICIAL_PARTY_HEADERS:
        return f"GCE-{sigla}"

    return sigla


def entity_type_for_sigla(sigla: str) -> str:
    if sigla in OFFICIAL_PARTY_HEADERS:
        return "party"

    if sigla.startswith("GCE-") or sigla.startswith("MOV") or sigla.startswith("M.A"):
        return "gce"

    if "." in sigla or "/" in sigla:
        return "coalition"

    return "party"


@dataclass(frozen=True)
class ParsedRow:
    row_no: int
    raw_code: str
    territory_code: str
    concelho: str | None
    freguesia: str | None
    office_code: str
    registered_voters: int
    voters: int
    blank_votes: int
    null_votes: int
    votes: list[tuple[str, str, int, int]]
    raw: dict[str, Any]


class Parser:
    FIXED_COLS = 8

    def __init__(self, path: Path, sheet_name: str | None = None):
        if path.suffix.lower() == ".xls":
            raise SystemExit(
                "Wrong file format. Convert .xls to .xlsx first with LibreOffice."
            )

        self.path = path
        self.workbook = load_workbook(path, read_only=True, data_only=True)
        self.sheet = (
            self.workbook[sheet_name]
            if sheet_name
            else self.workbook[self.workbook.sheetnames[0]]
        )
        self.sheet_name = self.sheet.title

        self.header_row = self.detect_header_row()
        self.data_start_row = self.header_row + 1

        self.group_headers = [
            clean_text(c.value) for c in self.sheet[self.header_row - 1]
        ]
        self.headers = [clean_text(c.value) for c in self.sheet[self.header_row]]

        self.last_coalition_labels_ordered: list[str] = []
        self.last_gce_labels_ordered: list[str] = []

        # Some CNE 2021 files have row 4 as:
        # CÓD, CONC, FREG, A, B.E., ...
        # but data rows still are:
        # CÓD, CONC, FREG, ÓRG, inscritos, votantes, brancos, nulos, party...
        # Force the first 8 headers to the operational schema.
        if len(self.headers) >= 8:
            first_three = [normalize_header_text(v) for v in self.headers[:3]]

            if first_three == ["COD", "CONC", "FREG"]:
                self.headers[0:8] = [
                    "CÓD",
                    "CONC",
                    "FREG",
                    "ÓRG",
                    "inscritos",
                    "votantes",
                    "brancos",
                    "nulos",
                ]

    def detect_header_row(self) -> int:
        for row_no in range(1, 30):
            row_values = [normalize_header_text(c.value) for c in self.sheet[row_no]]

            has_code = any(
                v in {"COD", "COD.", "CODIGO", "CODIGO.", "CÓD", "CÓD."}
                for v in row_values
            )
            has_conc = "CONC" in row_values
            has_freg = "FREG" in row_values

            has_office = any(
                v in {"ORG", "ORG.", "ORGAO", "ORGAO.", "ÓRG", "ÓRG."}
                for v in row_values
            )
            has_registered = any(v in {"INSC", "INSCRITOS"} for v in row_values)
            has_voters = any(v in {"VOT", "VOTANTES"} for v in row_values)

            # 2017/2025 standard: full fixed-column header row.
            if has_code and has_office and has_registered and has_voters:
                return row_no

            # 2021 original: row 4 only has CÓD/CONC/FREG, while row 3 has
            # ÓRG/INSC/VOT as group headers. Data rows still follow the same
            # fixed column order.
            if has_code and has_conc and has_freg:
                return row_no

        raise RuntimeError(
            f"Could not detect header row in sheet {self.sheet_name}. "
            "Expected a row containing code headers."
        )

    def result_columns(self) -> list[tuple[int, str]]:
        result_cols: list[tuple[int, str]] = []

        for idx, header in enumerate(self.headers, start=1):
            if idx <= self.FIXED_COLS or not header:
                continue

            is_placeholder = re.fullmatch(r"\[[A-Z]\]", header) is not None
            is_special_label_col = "SIGLA" in header.upper()

            if is_special_label_col:
                continue

            if header.startswith("[") and header.endswith("]") and not is_placeholder:
                continue

            result_cols.append((idx, header))

        return result_cols

    def is_column_group(self, idx: int, wanted: str) -> bool:
        header = self.headers[idx] if idx < len(self.headers) else None
        group_header = (
            self.group_headers[idx] if idx < len(self.group_headers) else None
        )

        header_norm = clean_text(header)
        group_norm = clean_text(group_header)

        if wanted == "coalition":
            return header_norm == "[SIGLA COL]" or group_norm == "SIGLAS COLIGAÇÕES"

        if wanted == "gce":
            return header_norm == "[SIGLA GCE]" or group_norm == "SIGLAS GCE"

        return False

    def row_coalition_labels(self, values: list[Any]) -> list[str]:
        labels: list[str] = []

        for idx, value in enumerate(values):
            header = self.headers[idx] if idx < len(self.headers) else None
            header_is_placeholder = (
                header is not None and re.fullmatch(r"\[[A-Z]\]", header) is not None
            )

            if self.is_column_group(idx, "coalition"):
                for label in parse_bracket_labels(value):
                    if label not in labels:
                        labels.append(label)

            elif self.is_column_group(idx, "gce"):
                continue

            elif idx >= self.FIXED_COLS and not header_is_placeholder:
                for label in parse_bracket_labels(value):
                    if looks_like_coalition_sigla(label) and label not in labels:
                        labels.append(label)

        return labels

    def row_gce_labels(self, values: list[Any]) -> list[str]:
        labels: list[str] = []

        for idx, value in enumerate(values):
            header = self.headers[idx] if idx < len(self.headers) else None
            header_is_placeholder = (
                header is not None and re.fullmatch(r"\[[A-Z]\]", header) is not None
            )

            if self.is_column_group(idx, "gce"):
                for label in parse_bracket_labels(value):
                    label = safe_gce_sigla(label)

                    if label not in labels:
                        labels.append(label)

            elif self.is_column_group(idx, "coalition"):
                continue

            elif idx >= self.FIXED_COLS and not header_is_placeholder:
                for label in parse_bracket_labels(value):
                    if not looks_like_coalition_sigla(label):
                        label = safe_gce_sigla(label)

                        if label not in labels:
                            labels.append(label)

        return labels

    def parse_rows(self) -> Iterable[ParsedRow]:
        result_columns = self.result_columns()

        for row_no, row in enumerate(
            self.sheet.iter_rows(min_row=self.data_start_row, values_only=True),
            start=self.data_start_row,
        ):
            values = list(row)

            raw_code = normalize_code(values[0] if len(values) > 0 else None)
            concelho = clean_text(values[1] if len(values) > 1 else None)
            freguesia = clean_text(values[2] if len(values) > 2 else None)
            office_code = clean_text(values[3] if len(values) > 3 else None)

            if not raw_code or not office_code:
                continue

            registered = as_int(values[4] if len(values) > 4 else None)
            voters = as_int(values[5] if len(values) > 5 else None)
            blank = as_int(values[6] if len(values) > 6 else None)
            null = as_int(values[7] if len(values) > 7 else None)

            if registered is None or voters is None or blank is None or null is None:
                continue

            territory_code = canonical_territory_code(
                raw_code,
                office_code,
                freguesia,
            )

            row_coalition_labels = self.row_coalition_labels(values)
            row_gce_labels = self.row_gce_labels(values)

            if row_coalition_labels:
                self.last_coalition_labels_ordered = row_coalition_labels

            if row_gce_labels:
                self.last_gce_labels_ordered = row_gce_labels

            coalition_labels_ordered = (
                row_coalition_labels
                if row_coalition_labels
                else self.last_coalition_labels_ordered
            )

            gce_labels_ordered = (
                row_gce_labels if row_gce_labels else self.last_gce_labels_ordered
            )

            votes: list[tuple[str, str, int, int]] = []
            coalition_votes_seen = 0
            gce_votes_seen = 0
            used_vote_columns: set[int] = set()

            for display_order, (col_idx, raw_sigla) in enumerate(
                result_columns,
                start=1,
            ):
                used_vote_columns.add(col_idx)

                vote_count = as_int(
                    values[col_idx - 1] if col_idx - 1 < len(values) else None
                )

                if vote_count is None:
                    continue

                if re.fullmatch(r"\[[A-Z]\]", raw_sigla):
                    placeholder_letter = raw_sigla.strip("[]")

                    if placeholder_letter in {"A", "B", "C"}:
                        if coalition_votes_seen < len(coalition_labels_ordered):
                            sigla = coalition_labels_ordered[coalition_votes_seen]
                        else:
                            if vote_count == 0:
                                coalition_votes_seen += 1
                                continue

                            sigla = (
                                f"SOURCE-COL-{placeholder_letter}-"
                                f"{territory_code}-{office_code}"
                            )

                        coalition_votes_seen += 1
                        entity_type = "coalition"

                    elif placeholder_letter in {"D", "E", "F", "G"}:
                        if gce_votes_seen < len(gce_labels_ordered):
                            sigla = gce_labels_ordered[gce_votes_seen]
                        else:
                            if vote_count == 0:
                                gce_votes_seen += 1
                                continue

                            sigla = (
                                f"SOURCE-GCE-{placeholder_letter}-"
                                f"{territory_code}-{office_code}"
                            )

                        gce_votes_seen += 1
                        entity_type = "gce"

                    else:
                        if vote_count == 0:
                            continue

                        sigla = (
                            f"SOURCE-OTHER-{placeholder_letter}-"
                            f"{territory_code}-{office_code}"
                        )
                        entity_type = "other"

                else:
                    sigla = raw_sigla
                    entity_type = entity_type_for_sigla(sigla)

                votes.append((sigla, entity_type, vote_count, display_order))

            raw = {
                str(self.headers[i] or f"col_{i + 1}"): values[i]
                for i in range(min(len(values), len(self.headers)))
            }

            yield ParsedRow(
                row_no=row_no,
                raw_code=raw_code,
                territory_code=territory_code,
                concelho=concelho,
                freguesia=freguesia,
                office_code=office_code,
                registered_voters=registered,
                voters=voters,
                blank_votes=blank,
                null_votes=null,
                votes=votes,
                raw=raw,
            )


def get_one(cur, sql: str, params: tuple[Any, ...]) -> Any:
    cur.execute(sql, params)
    row = cur.fetchone()

    if row is None:
        raise RuntimeError(f"Query returned no rows: {sql}")

    return row[0]


def build_territory_rows(
    parsed_rows: list[ParsedRow],
) -> list[tuple[str, str, str, str | None]]:
    territories: dict[str, tuple[str, str, str, str | None]] = {}

    territories["PT"] = ("PT", "country", "Portugal", None)

    for pr in parsed_rows:
        raw = pr.raw_code.zfill(6)

        district_code = raw[:2]
        municipality_code = raw[:4]
        parish_code = raw

        territories.setdefault(
            district_code,
            (
                district_code,
                "district",
                f"Distrito {district_code}",
                "PT",
            ),
        )

        # District/header placeholder rows such as 030000 must not create
        # fake municipalities like 0300.
        # But old CNE files may contain AF rows with xx0000 as real
        # election-source placeholder parish rows.
        if raw.endswith("0000"):
            if pr.office_code == "AF" and pr.freguesia:
                territories[parish_code] = (
                    parish_code,
                    "parish",
                    pr.freguesia,
                    district_code,
                )

            continue

        # Invalid municipality code xx00.
        # Do not create fake municipality 0300/0400/etc.
        if municipality_code.endswith("00"):
            # But if the row is AF and has a freguesia label, keep the exact
            # source code as an election-only parish placeholder. It will not
            # receive CAOP geometry, but the official election row can load.
            if pr.office_code == "AF" and pr.freguesia:
                territories[parish_code] = (
                    parish_code,
                    "parish",
                    pr.freguesia,
                    district_code,
                )

            continue

        territories[municipality_code] = (
            municipality_code,
            "municipality",
            pr.concelho or municipality_code,
            district_code,
        )

        # For AF rows, keep the exact source parish code, including xxxx00
        # placeholders. They are valid election-source rows, even if CAOP
        # cannot match them spatially.
        if pr.office_code == "AF" and pr.freguesia:
            territories[parish_code] = (
                parish_code,
                "parish",
                pr.freguesia,
                municipality_code,
            )

    level_order = {
        "country": 0,
        "district": 1,
        "municipality": 2,
        "parish": 3,
    }

    return sorted(
        territories.values(),
        key=lambda x: (level_order[x[1]], x[0]),
    )


def bulk_upsert_territories(
    cur,
    territories: list[tuple[str, str, str, str | None]],
) -> None:
    for code, level_code, name, parent_code in territories:
        cur.execute(
            """
            INSERT INTO op.territory (
                level_id,
                code,
                name,
                parent_id,
                normalized_name,
                geom,
                source_table,
                source_srid
            )
            SELECT
                tl.territory_level_id,
                %s,
                %s,
                parent.territory_id,
                lower(unaccent(%s)),
                NULL,
                'election_etl_fallback',
                NULL
            FROM op.territory_level tl
            LEFT JOIN op.territory parent
              ON parent.code = %s
            WHERE tl.code = %s
            ON CONFLICT (code)
            DO UPDATE SET
                name = COALESCE(op.territory.name, EXCLUDED.name),
                parent_id = COALESCE(op.territory.parent_id, EXCLUDED.parent_id),
                normalized_name = COALESCE(
                    op.territory.normalized_name,
                    EXCLUDED.normalized_name
                ),
                updated_at = now()
            """,
            (
                code,
                name or code,
                name or code,
                parent_code,
                level_code,
            ),
        )


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--database-url", default=os.getenv("DATABASE_URL"))
    ap.add_argument("--file", required=True, type=Path)
    ap.add_argument("--sheet")
    ap.add_argument("--election-code", default="AUTARQUICAS_2021")
    args = ap.parse_args()

    if not args.database_url:
        raise SystemExit(
            "DATABASE_URL missing. Set it in .env/export it or pass --database-url."
        )

    parser = Parser(args.file, args.sheet)
    file_hash = sha256_file(args.file)

    conn = psycopg2.connect(args.database_url)
    conn.autocommit = False

    rows_seen = 0
    rows_loaded = 0
    vote_cells = 0

    with conn, conn.cursor() as cur:
        election_id = get_one(
            cur,
            "SELECT election_id FROM op.election WHERE code = %s",
            (args.election_code,),
        )

        import_file_id = get_one(
            cur,
            """
            INSERT INTO op.import_file (
                election_id,
                file_path,
                sheet_name,
                file_hash
            )
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (election_id, file_hash, sheet_name)
            DO UPDATE SET imported_at = now()
            RETURNING import_file_id
            """,
            (
                election_id,
                str(args.file),
                parser.sheet_name,
                file_hash,
            ),
        )

        parsed_rows = list(parser.parse_rows())

        territories = build_territory_rows(parsed_rows)
        bulk_upsert_territories(cur, territories)

        for pr in parsed_rows:
            rows_seen += 1

            cur.execute(
                "SELECT op.save_turnout_result(%s,%s,%s,%s,%s,%s,%s,%s)",
                (
                    args.election_code,
                    pr.office_code,
                    pr.territory_code,
                    pr.registered_voters,
                    pr.voters,
                    pr.blank_votes,
                    pr.null_votes,
                    import_file_id,
                ),
            )

            for sigla, entity_type, votes, display_order in pr.votes:
                cur.execute(
                    "SELECT op.save_candidacy_vote_result(%s,%s,%s,%s,%s,%s,%s,%s)",
                    (
                        args.election_code,
                        pr.office_code,
                        pr.territory_code,
                        sigla,
                        entity_type,
                        votes,
                        display_order,
                        import_file_id,
                    ),
                )
                vote_cells += 1

            rows_loaded += 1
        cur.execute("CALL op.populate_seat_count();")
        cur.execute("CALL op.calculate_seat_results()")
        cur.execute("CALL wh.refresh_wh()")

    conn.close()

    print(
        json.dumps(
            {
                "file": str(args.file),
                "sheet": parser.sheet_name,
                "election_code": args.election_code,
                "territories_preloaded": len(territories),
                "rows_seen": rows_seen,
                "rows_loaded": rows_loaded,
                "vote_cells_loaded": vote_cells,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
